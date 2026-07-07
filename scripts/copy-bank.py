#!/usr/bin/env python3
# 스마트스토어 리뷰 엑셀 + 상품 문의(커머스API) → AI 카피 소스 뱅크
# 사용법:
#   python3 scripts/copy-bank.py                # ~/Downloads에서 최신 리뷰 엑셀 자동 탐색 (+Q&A)
#   python3 scripts/copy-bank.py 파일.xlsx      # 리뷰 파일 직접 지정 (+Q&A)
#   리뷰 엑셀이 없으면 Q&A만으로 실행됨 (네이버 커머스API는 리뷰 조회 미제공)
# 출력:
#   copy-bank/<제품>.md           — 제품별 카피 소스 (verbatim/상황/훅 등)
#   copy-bank/qna_망설임.md       — 문의 기반 구매 망설임 분석
#   Supabase settings oa_copy_bank_v1 — 대시보드에서 나중에 읽을 수 있게 저장
import base64, datetime, glob, json, os, re, sys, time, urllib.parse, urllib.request

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
env = {}
with open(os.path.join(ROOT, '.env.local')) as f:
    for line in f:
        m = re.match(r'([A-Z_]+)="?([^"\n]*)"?\s*$', line.strip())
        if m:
            env[m.group(1)] = m.group(2)
SUPA = env['NEXT_PUBLIC_SUPABASE_URL']
SUPA_KEY = env['NEXT_PUBLIC_SUPABASE_ANON_KEY']
GROQ_KEY = env.get('GROQ_API_KEY', '')
NC_ID = env.get('NAVER_COMMERCE_CLIENT_ID', '')
NC_SECRET = env.get('NAVER_COMMERCE_CLIENT_SECRET', '')

QNA_MONTHS = 6  # 문의 수집 기간


BEAUTY_ONLY = '--beauty' in sys.argv  # 이미용(드라이기/고데기/갈바닉/화장거울)만
BEAUTY_RE = re.compile(r'드라이|고데기|갈바닉|거울|미러', re.I)


def find_file():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    if args:
        return os.path.expanduser(args[0])
    pats = ['~/Downloads/*리뷰*.xlsx', '~/Downloads/*[Rr]eview*.xlsx']
    files = []
    for p in pats:
        files += glob.glob(os.path.expanduser(p))
    return max(files, key=os.path.getmtime) if files else None


# ── 네이버 커머스API: 상품 문의(Q&A) 수집 ──
def commerce_token():
    import bcrypt
    ts = str(int(time.time() * 1000))
    sign = base64.b64encode(bcrypt.hashpw(f'{NC_ID}_{ts}'.encode(), NC_SECRET.encode())).decode()
    body = urllib.parse.urlencode({'client_id': NC_ID, 'timestamp': ts,
                                   'grant_type': 'client_credentials', 'client_secret_sign': sign,
                                   'type': 'SELF'}).encode()
    req = urllib.request.Request('https://api.commerce.naver.com/external/v1/oauth2/token',
                                 data=body, headers={'Content-Type': 'application/x-www-form-urlencoded'},
                                 method='POST')
    return json.loads(urllib.request.urlopen(req, timeout=30).read())['access_token']


def fetch_qnas():
    tok = commerce_token()
    now = datetime.datetime.now()
    frm = (now - datetime.timedelta(days=QNA_MONTHS * 30)).strftime('%Y-%m-%dT00:00:00.000+09:00')
    to = now.strftime('%Y-%m-%dT%H:%M:%S.000+09:00')
    out, page = [], 1
    while True:
        q = urllib.parse.urlencode({'fromDate': frm, 'toDate': to, 'page': page, 'size': 100})  # 최대 100
        req = urllib.request.Request(f'https://api.commerce.naver.com/external/v1/contents/qnas?{q}',
                                     headers={'Authorization': f'Bearer {tok}'})
        for attempt in range(4):
            try:
                d = json.loads(urllib.request.urlopen(req, timeout=30).read())
                break
            except urllib.error.HTTPError as e:
                if e.code == 429 and attempt < 3:
                    time.sleep(2 * (attempt + 1))
                    continue
                raise
        time.sleep(0.4)  # rate limit 여유
        for c in d.get('contents', []):
            if c.get('question'):
                out.append({'product': c.get('productName', ''), 'question': c['question'].strip()[:300],
                            'date': (c.get('createDate') or '')[:10]})
        if d.get('last') or page >= d.get('totalPages', 1) or page > 50:
            break
        page += 1
    return out


def parse(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c or '') for c in next(it)]

    def col(*cands):
        for i, h in enumerate(header):
            if any(c in h for c in cands):
                return i
        return None

    ci_name = col('상품명', '제품명', '품명')
    ci_rating = col('구매자평점', '별점', '평점', '점수')
    ci_content = col('리뷰상세내용', '리뷰내용', '리뷰 내용', '내용', '후기')
    ci_option = col('옵션')
    if ci_content is None:
        sys.exit(f'리뷰 내용 컬럼을 못 찾았어요. 컬럼: {", ".join(header)}')

    rows = []
    for r in it:
        content = str(r[ci_content] or '').strip()
        if len(content) < 3:
            continue
        rating = 0
        if ci_rating is not None:
            rating = int(re.sub(r'[^0-9]', '', str(r[ci_rating] or '0'))[:1] or 0)
        rows.append({
            'product': str(r[ci_name] or '').strip() if ci_name is not None else os.path.basename(path),
            'rating': rating,
            'content': content[:500],
            'option': str(r[ci_option] or '').strip() if ci_option is not None else '',
        })
    return rows


def sample(reviews, limit, clip):
    """카피 소스용 샘플: 내용 긴 리뷰 우선 + 저평점(반박 소스) 포함"""
    neg = sorted([r for r in reviews if r['rating'] <= 2], key=lambda r: -len(r['content']))[:max(8, limit // 5)]
    mid = sorted([r for r in reviews if r['rating'] == 3], key=lambda r: -len(r['content']))[:5]
    pos = sorted([r for r in reviews if r['rating'] >= 4], key=lambda r: -len(r['content']))[:limit - len(neg) - len(mid)]
    return [{**r, 'content': r['content'][:clip]} for r in neg + pos + mid]


PROMPT = '''당신은 이미용 가전 브랜드 "오아"의 메타(페이스북/인스타) 광고 카피라이터입니다.
아래는 "{product}" 제품의 실제 고객 리뷰입니다.

{reviews}

목적: 광고 소재 제작에 바로 쓸 "카피 소스 뱅크"를 만드는 것. 마케터 언어가 아니라 고객이 실제 쓴 언어를 캐내세요.

아래 JSON으로만 답하세요. 다른 텍스트 금지:
{{
  "verbatims": [{{"text": "리뷰에서 거의 그대로 딴 광고에 쓸 만한 문장", "why": "왜 강력한지 1줄"}}, ...8~12개],
  "situations": [{{"scene": "구매/사용 상황 (예: 자취방 이사, 여행 준비)", "quote": "근거 리뷰 표현"}}, ...5~8개],
  "triggers": [{{"trigger": "구매 결정 계기 (예: 기존 제품 고장, 인스타에서 봄)", "count": 언급횟수추정}}, ...4~6개],
  "switching": [{{"from": "갈아타기 전 (예: 다이슨, 저가 드라이기)", "reason": "갈아탄 이유"}}, ...최대5개],
  "objections": [{{"complaint": "자주 나오는 불만/망설임", "counter": "소재에서 선제 대응할 카피"}}, ...3~5개],
  "hooks": [{{"hook": "첫 1초 훅 카피 (고객 언어 기반)", "appeal": "상황|가격|기능|비교 중 하나"}}, ...10개]
}}'''


QNA_PROMPT = '''당신은 이미용 가전 브랜드 "오아"의 메타 광고 카피라이터입니다.
아래는 스마트스토어 고객들이 구매 전후에 남긴 실제 상품 문의입니다.

{qnas}

문의는 "구매를 망설이게 하는 지점"의 원천 데이터입니다. 광고 소재/상세페이지에서 선제 대응할 수 있게 분석하세요.

아래 JSON으로만 답하세요. 다른 텍스트 금지:
{{
  "hesitations": [{{"question": "자주 나오는 망설임 (묶어서)", "count": 언급횟수추정, "counter": "소재/상세페이지에서 선제 대응할 카피", "product": "주로 어느 제품"}}, ...8~12개, count 많은 순],
  "unmet_needs": [{{"need": "문의에서 드러난 미충족 니즈/사용 시나리오", "idea": "소재 아이디어 1줄"}}, ...3~5개]
}}'''


def groq_chat(prompt, max_tokens=3500):
    req = urllib.request.Request(
        'https://api.groq.com/openai/v1/chat/completions',
        data=json.dumps({
            'model': 'llama-3.3-70b-versatile',
            'messages': [{'role': 'user', 'content': prompt}],
            'max_tokens': max_tokens,
            'temperature': 0.4,
        }).encode(),
        headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {GROQ_KEY}',
                 'User-Agent': 'oa-copy-bank/1.0'},  # Cloudflare가 파이썬 기본 UA 차단
        method='POST')
    with urllib.request.urlopen(req, timeout=120) as res:
        data = json.loads(res.read())
    raw = data['choices'][0]['message']['content']
    m = re.search(r'\{[\s\S]*\}', raw)
    if not m:
        raise ValueError('AI 응답 파싱 실패: ' + raw[:200])
    return json.loads(m.group(0))


def ai_chat(full_prompt, small_prompt):
    """1순위 claude CLI (품질 좋음, 입력 제한 넉넉) → 실패 시 Groq (작은 프롬프트, TPM 12k)"""
    import subprocess
    try:
        r = subprocess.run(['claude', '-p', '--model', 'sonnet'], input=full_prompt.encode(),
                           capture_output=True, timeout=600)
        if r.returncode == 0:
            m = re.search(r'\{[\s\S]*\}', r.stdout.decode())
            if m:
                return json.loads(m.group(0))
        print(f'  (claude CLI 실패 → Groq 폴백: {r.stderr.decode()[:100]})')
    except Exception as e:
        print(f'  (claude CLI 오류 → Groq 폴백: {e})')
    return groq_chat(small_prompt, max_tokens=2500)


def review_text(rows):
    return '\n'.join(f"[{i+1}] ★{r['rating']} {('['+r['option']+'] ') if r['option'] else ''}{r['content']}"
                     for i, r in enumerate(rows))


def analyze_reviews(product, reviews):
    full = PROMPT.format(product=product, reviews=review_text(sample(reviews, 130, 400)))
    small = PROMPT.format(product=product, reviews=review_text(sample(reviews, 50, 250)))
    return ai_chat(full, small)


def analyze_qnas(qnas):
    def lines(n, clip):
        return '\n'.join(f"({' '.join(q['product'].split()[:3])}) {q['question'][:clip]}" for q in qnas[:n])
    return ai_chat(QNA_PROMPT.format(qnas=lines(400, 200)), QNA_PROMPT.format(qnas=lines(120, 100)))


def qna_to_md(n, b):
    L = [f'# 상품 문의 기반 구매 망설임 분석', '', f'최근 {QNA_MONTHS}개월 문의 {n}개 분석 (네이버 커머스API)', '']
    L.append('## 망설임 TOP → 선제 반박 카피')
    for h in b.get('hesitations', []):
        L.append(f'- **{h["question"]}** (약 {h.get("count", "?")}회 · {h.get("product", "")})\n  - 대응: **{h.get("counter", "")}**')
    L.append('\n## 미충족 니즈 / 소재 아이디어')
    for u in b.get('unmet_needs', []):
        L.append(f'- {u["need"]} → {u.get("idea", "")}')
    return '\n'.join(L) + '\n'


def to_md(product, n_reviews, b):
    L = [f'# {product} — 카피 소스 뱅크', f'', f'리뷰 {n_reviews}개 분석 · 생성 스크립트: scripts/copy-bank.py', '']
    L.append('## 고객 표현 그대로 (verbatim)')
    for v in b.get('verbatims', []):
        L.append(f'- **"{v["text"]}"** — {v.get("why", "")}')
    L.append('\n## 구매/사용 상황')
    for s in b.get('situations', []):
        L.append(f'- {s["scene"]} — "{s.get("quote", "")}"')
    L.append('\n## 구매 트리거')
    for t in b.get('triggers', []):
        L.append(f'- {t["trigger"]} (약 {t.get("count", "?")}회)')
    if b.get('switching'):
        L.append('\n## 갈아타기 (경쟁 전환)')
        for s in b['switching']:
            L.append(f'- {s["from"]} → 오아: {s.get("reason", "")}')
    L.append('\n## 불만/망설임 → 선제 반박 카피')
    for o in b.get('objections', []):
        L.append(f'- 불만: {o["complaint"]}\n  - 반박: **{o.get("counter", "")}**')
    L.append('\n## 훅 후보 10')
    for i, h in enumerate(b.get('hooks', []), 1):
        L.append(f'{i}. [{h.get("appeal", "?")}] {h["hook"]}')
    return '\n'.join(L) + '\n'


def save_supabase(bank):
    # 기존 값과 병합 (리뷰만/Q&A만 실행해도 서로 안 지워지게)
    try:
        req = urllib.request.Request(
            f'{SUPA}/rest/v1/settings?key=eq.oa_copy_bank_v1&select=value',
            headers={'apikey': SUPA_KEY, 'Authorization': f'Bearer {SUPA_KEY}'})
        cur = json.loads(urllib.request.urlopen(req, timeout=30).read())
        old = cur[0]['value'] if cur and isinstance(cur[0].get('value'), dict) else {}
        bank = {**old, **bank}
    except Exception:
        pass
    body = json.dumps({'key': 'oa_copy_bank_v1', 'value': bank}, ensure_ascii=False).encode()
    req = urllib.request.Request(
        f'{SUPA}/rest/v1/settings', data=body,
        headers={'Content-Type': 'application/json', 'apikey': SUPA_KEY,
                 'Authorization': f'Bearer {SUPA_KEY}', 'Prefer': 'resolution=merge-duplicates'},
        method='POST')
    urllib.request.urlopen(req, timeout=30)


def main():
    if not GROQ_KEY:
        sys.exit('.env.local에 GROQ_API_KEY 없음')
    outdir = os.path.join(ROOT, 'copy-bank')
    os.makedirs(outdir, exist_ok=True)
    bank = {}

    # ── 1) 리뷰 엑셀 (있으면) ──
    path = find_file()
    if path:
        print(f'리뷰 파일: {path}')
        rows = parse(path)
        by_prod = {}
        for r in rows:
            by_prod.setdefault(r['product'], []).append(r)
        print(f'리뷰 {len(rows)}개 · 제품 {len(by_prod)}종')
        for product, reviews in by_prod.items():
            if len(reviews) < 10:
                print(f'  건너뜀 (리뷰 {len(reviews)}개 <10): {product}')
                continue
            print(f'  분석 중: {product} ({len(reviews)}개)...')
            try:
                b = analyze_reviews(product, reviews)
            except Exception as e:
                print(f'  ❌ 실패: {e}')
                continue
            bank[product] = {'total': len(reviews), **b}
            safe = re.sub(r'[^\w가-힣-]+', '_', product)[:60]
            md_path = os.path.join(outdir, f'{safe}.md')
            with open(md_path, 'w') as f:
                f.write(to_md(product, len(reviews), b))
            print(f'  ✅ {md_path}')
    else:
        print('리뷰 엑셀 없음 → Q&A만 분석 (리뷰는 스마트스토어 → 리뷰관리 → 엑셀 다운로드 후 ~/Downloads 저장)')

    # ── 2) 상품 문의 Q&A (커머스API) ──
    if NC_ID and NC_SECRET:
        print(f'상품 문의 수집 중 (최근 {QNA_MONTHS}개월)...')
        try:
            qnas = fetch_qnas()
            print(f'문의 {len(qnas)}개 수집')
            if BEAUTY_ONLY:
                qnas = [q for q in qnas if BEAUTY_RE.search(q['product'])]
                print(f'이미용 필터 후 {len(qnas)}개')
            if len(qnas) >= 20:
                qb = analyze_qnas(qnas)
                key = '_qna_beauty' if BEAUTY_ONLY else '_qna'
                bank[key] = {'total': len(qnas), 'months': QNA_MONTHS, **qb}
                md_path = os.path.join(outdir, 'qna_망설임_이미용.md' if BEAUTY_ONLY else 'qna_망설임.md')
                with open(md_path, 'w') as f:
                    f.write(qna_to_md(len(qnas), qb))
                print(f'  ✅ {md_path}')
        except Exception as e:
            print(f'  ❌ Q&A 분석 실패: {e}')
    else:
        print('NAVER_COMMERCE_CLIENT_ID/SECRET 없음 → Q&A 건너뜀')

    if bank:
        try:
            save_supabase(bank)
            print('✅ Supabase oa_copy_bank_v1 저장 완료 (대시보드 연동용)')
        except Exception as e:
            print(f'Supabase 저장 실패 (md 파일은 정상): {e}')
    print('완료')


if __name__ == '__main__':
    main()
