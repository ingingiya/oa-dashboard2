#!/usr/bin/env python3
# ~/Downloads의 쿠팡/지그재그 일일판매량 엑셀 → Supabase channel_daily_sales 업서트
# 쿠팡·지그재그는 ERP에 일괄 발주(사입)로 잡혀서, 이 파일의 실판매 수치를 별도 테이블에 쌓는다.
import glob, os, re, json, datetime, urllib.request

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
env = {}
with open(os.path.join(ROOT, '.env.local')) as f:
    for line in f:
        m = re.match(r'([A-Z_]+)="?([^"\n]*)"?\s*$', line.strip())
        if m:
            env[m.group(1)] = m.group(2)
SUPA = env['NEXT_PUBLIC_SUPABASE_URL']
KEY = env['NEXT_PUBLIC_SUPABASE_ANON_KEY']

DAYS = 35
today = datetime.date.today()
cutoff = today - datetime.timedelta(days=DAYS)
yesterday = (today - datetime.timedelta(days=1)).isoformat()

TG_CHAT = '8704535307'


def telegram(msg):
    try:
        tg = {}
        with open(os.path.expanduser('~/.claude/channels/telegram/.env')) as f:
            for line in f:
                if '=' in line:
                    k, v = line.strip().split('=', 1)
                    tg[k] = v.strip('"')
        token = tg.get('TELEGRAM_BOT_TOKEN')
        if not token:
            return
        req = urllib.request.Request(
            f'https://api.telegram.org/bot{token}/sendMessage',
            data=json.dumps({'chat_id': TG_CHAT, 'text': msg}).encode(),
            headers={'Content-Type': 'application/json'}, method='POST')
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        print('텔레그램 알림 실패:', e)


def latest(pattern):
    files = glob.glob(os.path.expanduser(f'~/Downloads/{pattern}'))
    return max(files, key=os.path.getmtime) if files else None


def date_cols(header):
    out = {}
    for i, c in enumerate(header):
        if isinstance(c, datetime.datetime) and cutoff <= c.date() <= today:
            out[i] = c.date().isoformat()
    return out


rows = {}  # (channel,name,date) -> row dict


def add(channel, category, name, date, qty):
    try:
        q = float(qty)
    except (TypeError, ValueError):
        return
    if q <= 0:
        return
    rows[(channel, name, date)] = {
        'channel': channel, 'category': category,
        'name': name, 'date': date, 'qty': q,
    }


# ── 쿠팡: 'N월 출고' 시트 (출고 = 실제 소비자 판매)
cp = latest('쿠팡 일일판매량*.xlsx')
if cp:
    wb = openpyxl.load_workbook(cp, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        if not sheet.endswith('출고'):
            continue
        it = wb[sheet].iter_rows(values_only=True)
        header = None
        for r in it:
            if r and r[0] == '담당자':
                header = list(r)
                break
        if not header:
            continue
        dates = date_cols(header)
        if not dates:
            continue
        name_i = header.index('SKU명') if 'SKU명' in header else 7
        cat_i = header.index('품목') if '품목' in header else 2
        for r in it:
            name = r[name_i] if len(r) > name_i else None
            if not name or not str(name).strip():
                continue
            for i, d in dates.items():
                if i < len(r) and r[i] is not None:
                    add('쿠팡', str(r[cat_i] or '').strip(), str(name).strip(), d, r[i])
    print(f'쿠팡 파싱: {os.path.basename(cp)}')
else:
    print('쿠팡 파일 없음')

# ── 지그재그: 월별 시트, 헤더에 '제품명'
zz = latest('지그재그(직진)_일일판매량*.xlsx')
if zz:
    wb = openpyxl.load_workbook(zz, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        it = wb[sheet].iter_rows(values_only=True)
        header = None
        for r in it:
            if r and '제품명' in [str(c) for c in r if c]:
                header = list(r)
                break
        if not header:
            continue
        dates = date_cols(header)
        if not dates:
            continue
        name_i = header.index('제품명')
        for r in it:
            name = r[name_i] if len(r) > name_i else None
            if not name or not str(name).strip():
                continue
            for i, d in dates.items():
                if i < len(r) and r[i] is not None:
                    add('지그재그', '', str(name).strip(), d, r[i])
    print(f'지그재그 파싱: {os.path.basename(zz)}')
else:
    print('지그재그 파일 없음')

# ── 어제 데이터 존재 여부 체크 (없으면 텔레그램 알림)
missing = []
for ch, label in [('쿠팡', cp), ('지그재그', zz)]:
    if not label:
        missing.append(f'{ch}: 파일 없음')
    elif not any(k[0] == ch and k[2] == yesterday for k in rows):
        missing.append(f'{ch}: 어제({yesterday}) 데이터 없음 — 최신 파일인지 확인 필요')
if missing:
    print('경고:', '; '.join(missing))
    if '--dry-run' not in os.sys.argv:
        telegram('⚠️ 실판매 동기화 경고\n' + '\n'.join(missing)
                 + '\n\n네이버웍스 메일 첨부파일을 ~/Downloads에 저장해주세요.')

# ── Supabase 업서트
data = list(rows.values())
print(f'업서트 대상: {len(data)}행 ({cutoff} ~ {today})')
if data and '--dry-run' not in os.sys.argv:
    for i in range(0, len(data), 500):
        batch = data[i:i + 500]
        req = urllib.request.Request(
            f'{SUPA}/rest/v1/channel_daily_sales?on_conflict=channel,name,date',
            data=json.dumps(batch).encode(),
            headers={
                'apikey': KEY, 'Authorization': f'Bearer {KEY}',
                'Content-Type': 'application/json',
                'Prefer': 'resolution=merge-duplicates,return=minimal',
            }, method='POST')
        try:
            urllib.request.urlopen(req)
        except urllib.error.HTTPError as e:
            err = e.read().decode()[:300]
            print('업서트 실패:', err)
            telegram(f'❌ 실판매 동기화 실패 (Supabase 업서트 오류)\n{err}')
            raise
    print('업서트 완료')
